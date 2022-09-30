from pyxlsb import open_workbook
import xlrd
import openpyxl
import os
import sys
import optparse
import requests
import re

URLHAUS_API = "https://urlhaus-api.abuse.ch/v1/url/"
regex_uri = "https?:\/\/[a-zA-Z0-9\.\/\-\:]{5,}\.\w{1,}"

def setup_args():
    parser = optparse.OptionParser()

    parser.add_option('-d', '--directory',
    action="store", dest="directory",
    help="The folder that contains your Dridex docs", default=".")

    return parser.parse_args()

def extract_downloader_links(file_path):
  
    print("[*] Working file: " + file_path)
    #xls = open_workbook(file_path)
    #xls = load_workbook(file_path, read_only=True)
    try:
        xls = xlrd.open_workbook(file_path)
    except: 
        return
    print(xls.sheet_names)

#    xls = openpyxl.load_workbook(file_path)
    sheetnames = xls.sheet_names()
    obf = []


    print("=========================")
    for sheet in sheetnames:
        print(sheet)
        s = xls.sheet_by_name(sheet)
        #obf.append(s)
        #continue

        ##################################
        # template code found online - needed for transferring functionality to xlrd API
        #   - could not get execution of provided URL script or from HP Threat Research script
        rows = s.nrows - 1
        cells = s.ncols - 1
        curr_row = -1
        while curr_row < rows:
            curr_row += 1
            row = s.row(curr_row)
            curr_cell = -1
            while curr_cell < cells:
                curr_cell += 1
                cell_type = s.cell_type(curr_row, curr_cell)
                cell_value = s.cell_value(curr_row, curr_cell)
        # end template code from online
        #
        ########################################
                if cell_type != 0:
                    obf.append(cell_value)
        
    urls = []
    script = ""
    tmp_urls = ""

    if len(obf) == 2:
        for y in range(len(obf[1])):
            tmp_urls = tmp_urls + chr(ord(obf[1][y]) + int(obf[0][y]))

        if tmp_urls:
            url_string = tmp_urls.split("RSab")
            for url in url_string[0].split("E,"):
                urls.append("https://" + url)
    else:
        for y in range(len(obf)):
            if isinstance(obf[y], float):
                script = script + chr(int(obf[y]))

        if script:
            tmp_urls = re.findall(regex_uri, script)
            for url in tmp_urls:
                urls.append(url)

    print("[*] Found " + str(len(urls)) + " URLs")
    for url in urls:
        print("[$] Found - " + url)
        data = {'url' : url}
        response = requests.post(URLHAUS_API, data)

        response = response.json()

        if response['larted'] == "true":
            print("\t[!] URL reported before")
            print("\t[!] Malware Families:")
            for i in response["payloads"]:
                print("\t\t" + i['signature'])
        else:
            print("\t[!] URL is new")
    


def main(argv):

    options, args = setup_args() 

    for file_name in os.listdir(options.directory):
        extract_downloader_links(options.directory + file_name)

if __name__ == '__main__':
	main(sys.argv[1:])
